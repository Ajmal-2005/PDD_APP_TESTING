import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_security_excel_reports():
    FONT_FAMILY = "Segoe UI"
    
    # Header & Accent Fills
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name=FONT_FAMILY, size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name=FONT_FAMILY, size=10, italic=True, color="4B5563")
    
    crit_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    crit_font = Font(name=FONT_FAMILY, size=10, bold=True, color="991B1B")
    
    high_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    high_font = Font(name=FONT_FAMILY, size=10, bold=True, color="C2410C")
    
    med_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    med_font = Font(name=FONT_FAMILY, size=10, bold=True, color="92400E")
    
    low_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    low_font = Font(name=FONT_FAMILY, size=10, bold=True, color="075985")
    
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    pass_font = Font(name=FONT_FAMILY, size=10, bold=True, color="166534")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # ---------------------------------------------------------
    # DATA DEFINITIONS
    # ---------------------------------------------------------
    findings = [
        [
            "SEC-001", "Hardcoded API Keys in Repository & Environment Config", "Critical", 
            "Secrets / Sensitive Data Exposure", ".env.local", "/api/weather",
            "Hardcoded OpenWeather API key (0da720e0041a9a84...) and Firebase Web API key present in .env.local file committed to repository.",
            "Attacker clones repository or inspects git history, extracts API keys, and abuses third-party quotas or unauthorized services.",
            "Financial cost, API rate limiting, unauthorized service utilization.",
            "Move all API keys to environment secrets/vault and remove secret values from version control history.", "Open"
        ],
        [
            "SEC-002", "Unauthenticated Edge Server-Side Proxy Endpoint", "High",
            "Broken Authentication / Authorization", "src/app/api/weather/route.ts", "/api/weather?lat={lat}&lon={lon}",
            "The /api/weather GET endpoint accepts latitude and longitude without validating user session tokens or Firebase Auth JWT.",
            "Unauthenticated remote user or bot queries /api/weather repeatedly, proxying requests to OpenWeather API and exhausting rate limits.",
            "Denial of Service (DoS) on weather features, API key quota exhaustion.",
            "Enforce Firebase ID Token verification via middleware or request handler before forwarding proxy request.", "Open"
        ],
        [
            "SEC-003", "Client-Side Trust & Local State Mutation for Scan Data", "High",
            "Business Logic / Client-Side Trust", "src/lib/classifier.ts", "Client-Side TFJS Classifier",
            "Plant disease identification and confidence scoring run entirely in client browser context using TensorFlow.js with zero server validation.",
            "Malicious user modifies client-side JavaScript or local Dexie DB records, forging diagnostic results, confidence scores, or advisory PDFs.",
            "Tampered agricultural diagnostic records, false disease reports, inaccurate field analytics.",
            "Implement server-side or Cloud Function re-verification for official diagnostic logging and PDF generation.", "Open"
        ],
        [
            "SEC-004", "Missing Strict CSP & Security Response Headers", "Medium",
            "Security Misconfiguration", "next.config.mjs", "Global App Response Headers",
            "Missing HTTP Security Headers including Content-Security-Policy (CSP), Strict-Transport-Security (HSTS), X-Frame-Options, and X-Content-Type-Options.",
            "Attacker performs Cross-Site Scripting (XSS) or clickjacking attacks by embedding the web app in malicious iframes.",
            "Client-side session compromise, UI redressing, data theft.",
            "Add security headers in next.config.mjs or middleware.ts (CSP, HSTS, X-Frame-Options: DENY, X-Content-Type-Options: nosniff).", "Open"
        ],
        [
            "SEC-005", "Client-Side IndexedDB Data Exposure & Missing Encryption", "Medium",
            "Sensitive Data Exposure", "src/lib/db.ts", "IndexedDB (Dexie)",
            "Farmer scan history, location coordinates, notes, and disease diagnostic logs stored in plain browser IndexedDB without encryption at rest.",
            "Shared mobile/desktop device user or malicious browser extension reads local Dexie database store.",
            "Unauthorized access to local farm diagnostic history and GPS locations.",
            "Apply client-side payload encryption for sensitive Dexie DB fields or use browser Web Crypto API.", "Open"
        ],
        [
            "SEC-006", "Known Dependency Vulnerabilities in Next.js & SWC", "Low",
            "Dependency Vulnerabilities", "package.json", "NPM Dependencies",
            "Outdated development and core packages in package.json (Next.js 14.2.35, TensorFlow.js alpha dependencies).",
            "Exploitation of public CVEs in framework runtime or sub-dependencies.",
            "Potential Denial of Service or remote code execution depending on sub-dependency advisory.",
            "Upgrade Next.js to latest stable patch and update package-lock.json using npm audit fix.", "Open"
        ]
    ]

    endpoints = [
        ["/api/weather", "GET", "No", "None (Public)", "src/app/api/weather/route.ts", "Proxies OpenWeather API requests using server secret key"],
        ["/login", "GET/POST", "No", "Public", "src/app/login/page.tsx", "User authentication interface using Firebase Client Auth"],
        ["/register", "GET/POST", "No", "Public", "src/app/register/page.tsx", "New farmer account creation screen"],
        ["/dashboard", "GET", "Yes (Client)", "Authenticated User", "src/app/(app)/dashboard/page.tsx", "Farmer overview, weather overview, recent scans"],
        ["/scan", "GET/POST", "Yes (Client)", "Authenticated User", "src/app/(app)/scan/page.tsx", "Leaf image upload & TensorFlow.js plant disease identification"],
        ["/history", "GET", "Yes (Client)", "Authenticated User", "src/app/(app)/history/page.tsx", "Local Dexie IndexedDB scan history & filtering"],
        ["/library", "GET", "No", "Public", "src/app/(app)/library/page.tsx", "Agronomic disease catalog & symptom reference"],
        ["/analytics", "GET", "Yes (Client)", "Authenticated User", "src/app/(app)/analytics/page.tsx", "Field diagnostic metrics & crop health charts"],
        ["/profile", "GET/POST", "Yes (Client)", "Authenticated User", "src/app/(app)/profile/page.tsx", "Farmer profile settings & Firebase user details"],
        ["/settings", "GET/POST", "Yes (Client)", "Authenticated User", "src/app/(app)/settings/page.tsx", "App preferences, language switcher, cache management"]
    ]

    dependencies = [
        ["next", "14.2.35", "High", "CVE-2024-34351 / Sub-dependency advisories", "Upgrade to Next.js >= 14.2.23 / 15.x"],
        ["@tensorflow/tfjs-tflite", "0.0.1-alpha.10", "Medium", "Alpha pre-release security & stability risks", "Migrate to stable TFJS WebGL/CPU pipeline"],
        ["firebase", "11.1.0", "Low", "No high CVEs (Requires strict Firestore rules)", "Maintain Firestore security rules"],
        ["dexie", "4.0.10", "Low", "No active CVEs", "Keep updated"],
        ["jspdf", "2.5.2", "Low", "DOMPurify sub-dependency advisory", "Update sub-dependencies via npm audit fix"]
    ]

    risks = [
        ["Critical", 1, "Hardcoded OpenWeather & Firebase API Secrets in Repository"],
        ["High", 2, "Unauthenticated Weather Proxy Endpoint & Client-Side Trust Assumption"],
        ["Medium", 2, "Missing Security Response Headers (CSP/HSTS) & Unencrypted Local Storage"],
        ["Low", 1, "Known Dependency Security Advisories in Web Framework"]
    ]

    # ---------------------------------------------------------
    # WORKBOOK 1: endpoint-inventory.xlsx
    # ---------------------------------------------------------
    wb_ep = openpyxl.Workbook()
    ws_ep1 = wb_ep.active
    ws_ep1.title = "Endpoint Inventory"
    ws_ep1.views.sheetView[0].showGridLines = True
    
    ws_ep1["A1"] = "AgroVision Web Application - Endpoint Inventory"
    ws_ep1["A1"].font = title_font
    ws_ep1["A2"] = "Complete discovery of public and private Next.js App Router API & Page Routes"
    ws_ep1["A2"].font = subtitle_font

    ep_headers = ["Endpoint URL", "HTTP Method", "Authentication Required", "Expected Roles", "Controller / File Path", "Description & Logic"]
    for col_idx, h in enumerate(ep_headers, start=1):
        cell = ws_ep1.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(endpoints, start=5):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_ep1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=FONT_FAMILY, size=10)
            cell.border = thin_border
            if col_idx in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 3:
                if val == "Yes (Client)":
                    cell.fill = pass_fill
                    cell.font = pass_font
                else:
                    cell.fill = med_fill
                    cell.font = med_font

    wb_ep.save("Vulnerability Test Results/endpoint-inventory.xlsx")

    # ---------------------------------------------------------
    # WORKBOOK 2: findings.xlsx (Multi-Sheet Workbook)
    # ---------------------------------------------------------
    wb_f = openpyxl.Workbook()

    # SHEET 1: Security Findings
    ws_s1 = wb_f.active
    ws_s1.title = "Security Findings"
    ws_s1.views.sheetView[0].showGridLines = True

    ws_s1["A1"] = "AgroVision Web Application - Static & Dynamic SAST/DAST Security Findings"
    ws_s1["A1"].font = title_font
    ws_s1["A2"] = "Comprehensive Security Vulnerability Audit Log"
    ws_s1["A2"].font = subtitle_font

    f_headers = ["Finding ID", "Vulnerability Title", "Severity", "Vulnerability Type", "File Path", "Endpoint", "Description", "Exploitation Scenario", "Impact", "Recommended Fix", "Status"]
    for col_idx, h in enumerate(f_headers, start=1):
        cell = ws_s1.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(findings, start=5):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_s1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=FONT_FAMILY, size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col_idx == 3: # Severity
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "Critical":
                    cell.fill = crit_fill
                    cell.font = crit_font
                elif val == "High":
                    cell.fill = high_fill
                    cell.font = high_font
                elif val == "Medium":
                    cell.fill = med_fill
                    cell.font = med_font
                else:
                    cell.fill = low_fill
                    cell.font = low_font
            elif col_idx in [1, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # SHEET 2: Endpoint Inventory
    ws_s2 = wb_f.create_sheet(title="Endpoint Inventory")
    ws_s2.views.sheetView[0].showGridLines = True
    ws_s2["A1"] = "AgroVision Web Application - Endpoint Inventory"
    ws_s2["A1"].font = title_font
    
    for col_idx, h in enumerate(ep_headers, start=1):
        cell = ws_s2.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(endpoints, start=4):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_s2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=FONT_FAMILY, size=10)
            cell.border = thin_border

    # SHEET 3: Dependency Vulnerabilities
    ws_s3 = wb_f.create_sheet(title="Dependency Vulnerabilities")
    ws_s3.views.sheetView[0].showGridLines = True
    ws_s3["A1"] = "Software Composition Analysis (SCA) - Dependency Vulnerabilities"
    ws_s3["A1"].font = title_font

    dep_headers = ["Package Name", "Installed Version", "Severity", "CVE / Advisory Details", "Remediation Action"]
    for col_idx, h in enumerate(dep_headers, start=1):
        cell = ws_s3.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(dependencies, start=4):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_s3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=FONT_FAMILY, size=10)
            cell.border = thin_border
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "High":
                    cell.fill = high_fill
                    cell.font = high_font
                elif val == "Medium":
                    cell.fill = med_fill
                    cell.font = med_font
                else:
                    cell.fill = low_fill
                    cell.font = low_font

    # SHEET 4: Risk Summary
    ws_s4 = wb_f.create_sheet(title="Risk Summary")
    ws_s4.views.sheetView[0].showGridLines = True
    ws_s4["A1"] = "Executive Security Risk Summary"
    ws_s4["A1"].font = title_font

    risk_headers = ["Severity Level", "Finding Count", "Primary Risk Description"]
    for col_idx, h in enumerate(risk_headers, start=1):
        cell = ws_s4.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(risks, start=4):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_s4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=FONT_FAMILY, size=10)
            cell.border = thin_border
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit column widths across all sheets
    for ws in wb_f.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and cell.row > 2:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 50)

    wb_f.save("Vulnerability Test Results/findings.xlsx")
    print("Successfully generated security report Excel workbooks in 'Vulnerability Test Results/'!")

if __name__ == "__main__":
    generate_security_excel_reports()
